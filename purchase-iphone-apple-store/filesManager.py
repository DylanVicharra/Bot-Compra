import pandas as pd 
import openpyxl as op
import glob
import json
from .errors import ExcelFileNoFound, ExcelContentNoSupported, JsonFileNoFound

class FilesHandler: 

    acceptedColumnsNames = []
    sheetName = None
    errorProduct = []

    def eliminationOfNulls(self, excelFile):
        excelFile = excelFile.dropna(axis=0, subset=['MODELO'])
        excelFile = excelFile.dropna(axis=0, subset=['OPERADOR'])
        excelFile = excelFile.dropna(axis=0, subset=['PASSWORD'])
        excelFile = excelFile.dropna(axis=0, subset=['USER'])
        excelFile = excelFile.dropna(axis=0, subset=['CANTIDAD'])
        return excelFile

    def getDriverPurchase(self, excelFile, sheetName):
        driverSheet = pd.read_excel(f'{excelFile}', sheet_name=sheetName, engine='openpyxl', header=1)

        driverInfo = {}
        driverInfo['name'] = driverSheet.iloc[0]['Nombre']
        driverInfo['lastName'] = driverSheet.iloc[0]['Apellido']

        return driverInfo


    def readJson(self, jsonFile):
        try:
            with open(jsonFile, encoding='utf-8') as jsonFileFound:
                return json.load(jsonFileFound)
        except:
            raise JsonFileNoFound(f'No se encontro un json con el nombre {jsonFile}')

    def getExcelFile(self, folderPath):
        # Return first excel file
        excelFile = glob.glob(f'{folderPath}/*.xlsx')
        if excelFile: 
            return excelFile[0]
        raise ExcelFileNoFound(f'No se encontro ningun archivo excel en {folderPath}')

    def writeExcelFile(self, excelFile, product):
        sheet = excelFile.get_sheet_by_name(self.sheetName)
        # last row of sheet
        lastRow = sheet.max_row
        # Write product information
        sheet.cell(row = lastRow+1, column = 1).hyperlink = product.numberOrder['link']
        sheet.cell(row = lastRow+1, column = 1).value = product.numberOrder['number']
        sheet.cell(row = lastRow+1, column = 1).style = 'Hyperlink'
        sheet.cell(row = lastRow+1, column = 2).value = product.productName
        sheet.cell(row = lastRow+1, column = 3).value = product.purchaseInformation.store["nombre"]

    def openOrCreateExcelFile(self, folderPath, excelFileName):
        try: 
            excelFile = op.load_workbook(f'{folderPath}\\{excelFileName}')
            return excelFile
        except:
            # Create a new excel file
            excelFile = op.Workbook()
            excelFile.worksheets[0].title = self.sheetName
            sheet = excelFile.get_sheet_by_name(self.sheetName)
            # Headers
            sheet.cell(row = 1, column = 1).value = "Nº ORDER"
            sheet.cell(row = 1, column = 2).value = "PRODUCT"
            sheet.cell(row = 1, column = 3).value = "STORE"
            # Cells widht
            sheet.column_dimensions['A'].width = 40
            sheet.column_dimensions['B'].width = 40
            sheet.column_dimensions['C'].width = 40
            return excelFile

    def readExcelFile(self, excelFile, sheetName):
        productsList = pd.read_excel(f'{excelFile}', sheet_name=sheetName, engine='openpyxl', header=0)
        productsList = self.eliminationOfNulls(productsList)
        purchaseProduct = []
        
        if list(productsList.columns.values):
            for i in productsList.index:
                try:
                    purchaseProduct.append(productsList.iloc[i])
                    #print("Value: " + purchaseProduct[i]['MODELO'])
                except Exception as ex:
                    errorMessage = '\n'.join(map(str, ex.args)).rstrip()
                    errProduct = {}
                    errProduct['errorName'] = errorMessage
                    errProduct['product'] = productsList.iloc[i]['MODELO'].lower()
                    self.errorProduct.append(errProduct)
            return purchaseProduct
        raise ExcelContentNoSupported('El contenido del archivo no es el adecuado.')

    def saveExcelFile(self, excelFile, excelFileName, folderPath):
        excelFile.save(f'{folderPath}\\{excelFileName}')
