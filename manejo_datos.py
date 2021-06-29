import json
from os import path
from pathlib import Path
from datetime import date
import openpyxl as op
import pandas as pd
from producto import Producto

src_path = Path(__file__).parent
data_path = src_path / 'data'


# Lectura de archivos json
def read_json(archivo):
    with open(archivo, encoding='utf-8') as archivo_json:
        return json.load(archivo_json)

# Verifica si existe el archivo excel 
def existe_archivo_excel(nombre_archivo):
    if path.exists(f'{nombre_archivo}.xlsx'):
        return True
    else: 
        raise Exception(f"Archivo no encontrado en la carpeta ./bot_compra4.1/{nombre_archivo}'") # Mal dicho pero ver si es correcto poner en otra carpeta distinta

# Verifica el modelo del producto
def verificacion_modelo(producto):
    lista_productos = read_json(data_path / 'iphones.json')
    for iphone in lista_productos:
        if producto == iphone:
            return lista_productos[iphone]
    raise Exception("No se encontro el modelo seleccionado en el archivo iphones.json") # de momento, ver para despues (colocar return False)

# Verifica el apple store dentro del archivo stores.json
def verificacion_apple_store(apple_store):
    lista_stores = read_json(data_path / 'stores.json')
    for store in lista_stores:
        if apple_store == lista_stores[store]["nombre"]:
            return lista_stores[store] 
    raise Exception("No se encontro la tienda seleccionada en el archivo stores.json") # de momento, ver para despues (colocar return False)

# PRUEBA
def listado_tiendas_opcionales(tiendas_excel, fila):
    
    tiendas_opcionales = {}

    for i in range(1, 4):
        # Con 'nan' se refiere que es vacio, lo cual no quiero agregar
        if str(tiendas_excel.loc[fila,f'OPCIONAL {i}']) != "nan":
            tiendas_opcionales[f'{i}'] = verificacion_apple_store(str(tiendas_excel.loc[fila, f'OPCIONAL {i}']))

    return tiendas_opcionales
    

def crear_archivo(nombre_archivo):
    if path.exists(f'{date.today()}-{nombre_archivo}'):
        archivo = op.load_workbook(f'{date.today()}-{nombre_archivo}')
        return archivo
    else: 
        # Se crea un nuevo archivo 
        archivo = op.Workbook()
        # Se renombra la primera hoja
        archivo.worksheets[0].title = "Ordenes"
        # Me muevo a la hoja
        hoja = archivo.get_sheet_by_name("Ordenes")
        # Tamaños predeterminados dados
        hoja.column_dimensions['A'].width = 25
        hoja.column_dimensions['B'].width = 40
        hoja.column_dimensions['C'].width = 30
        return archivo

def limpieza_datos(archivo_excel):
    # Se elimina las filas vacias que no contengan la informacion requerida evitando errores
    archivo_excel = archivo_excel.dropna(axis=0, subset=['MODELO'])
    archivo_excel = archivo_excel.dropna(axis=0, subset=['OPERADOR'])
    archivo_excel = archivo_excel.dropna(axis=0, subset=['PASSWORD'])
    archivo_excel = archivo_excel.dropna(axis=0, subset=['USER'])
    archivo_excel = archivo_excel.dropna(axis=0, subset=['CANTIDAD'])
    return archivo_excel

# Se elimina la fila y columna
def escribir_fila_excel(archivo_a_modificar, hoja, orden, nombre_producto, store):
    #Busco la hoja donde tengo que modificar 
    hoja_a_modificar = archivo_a_modificar.get_sheet_by_name(hoja)
    #Selecciono la ultima fila para asi no tener que guardar en un lugar especifico
    ultima_fila = hoja_a_modificar.max_row
    #Escribo en la columna uno que es el numero de orden:
    hoja_a_modificar.cell(row = ultima_fila+1, column = 1).hyperlink = orden["link"] #Hay que ver como guardar el orden
    hoja_a_modificar.cell(row = ultima_fila+1, column = 1).value = orden["nombre"]
    hoja_a_modificar.cell(row = ultima_fila+1, column = 1).style = 'Hyperlink'
    #hoja_a_modificar.cell(row = ultima_fila+1, column = 1).font = Font(underline='single')
    #Escribo en la columna dos que es el nombre del producto comprado:
    hoja_a_modificar.cell(row = ultima_fila+1, column = 2).value = nombre_producto
    #Escribo en la columna tres que es el nombre de la tienda que se va retirar, en el caso que sea de retiro 
    hoja_a_modificar.cell(row = ultima_fila+1, column = 3).value = store

# prueba de algo

def lectura_lista_compra(nombre_archivo, hoja):

    try:
        lista_compra = limpieza_datos(pd.read_excel(f'{nombre_archivo}.xlsx', sheet_name=hoja, engine='openpyxl'))
        print(f"Limpieza de datos del archivo excel '{nombre_archivo}' ... ")
    except:
        print(f"No se encuentra la hoja propuesta en el archivo excel {nombre_archivo}." + 
            "\n" + "Revise el archivo excel, renombre la hoja o cambie el valor de la variable 'hoja_excel'" + 
            "\n" + "Finalizando..." + '\n')
        exit(1)

    modelos_a_comprar = {}
    modelos_error = []

    for i in lista_compra.index:
        try:
            dic = {}
            dic['nombre'] = str(lista_compra.loc[i,"MODELO"])
            dic['objeto'] = Producto(*verificacion_modelo(str(lista_compra.loc[i,"MODELO"])),
                                     str(lista_compra.loc[i,"OPERADOR"]).lower(),
                                     int(lista_compra.loc[i,"CANTIDAD"]),
                                     str(lista_compra.loc[i,"USER"]),
                                     str(lista_compra.loc[i,"PASSWORD"]),
                                     verificacion_apple_store(str(lista_compra.loc[i,"STORE"])),
                                     listado_tiendas_opcionales(lista_compra, i)  
                                    )

            modelos_a_comprar[i] = dic

        except Exception as ex:
            error_message = '\n'.join(map(str, ex.args)).rstrip()
            modelos_error.append(f'* {str(lista_compra.loc[i,"MODELO"])} -> {error_message}')

    if modelos_error:
        print('\n' + ' - Error en archivos de datos: ')
        for error in modelos_error:
            print(error)

    return modelos_a_comprar


